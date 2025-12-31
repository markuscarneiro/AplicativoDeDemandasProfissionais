"""
Views do sistema de gestão de demandas.

Este módulo contém todas as views responsáveis por:
- Dashboard e estatísticas
- CRUD de demandas
- Matriz de Eisenhower
- Exportação de relatórios (Excel/PDF)
- Gestão de tags e comentários
- Sistema de notificações
"""

from datetime import datetime, timedelta
from typing import Any

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Count, F, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    UpdateView,
)

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .forms import AnexoForm, BuscaFiltroForm, ComentarioForm, DemandaForm
from .models import AnexoArquivo, Comentario, Demanda, HistoricoAlteracao, Tag


# =============================================================================
# CONSTANTES
# =============================================================================

DAYS_ALERT_THRESHOLD = 7  # Dias para alerta de prazo próximo
DEMANDAS_PER_PAGE = 20


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def dashboard(request):
    """
    View principal do sistema - Dashboard.
    
    Exibe estatísticas completas das demandas, alertas de prazo,
    distribuição por status e métricas de desempenho.
    """
    hoje = timezone.now().date()
    
    # Estatísticas por status
    status_counts = _get_status_counts()
    
    # Demandas atrasadas (prazo vencido e não concluídas)
    demandas_atrasadas = Demanda.objects.filter(
        data_prazo__lt=hoje
    ).exclude(
        status__in=[Demanda.Status.CONCLUIDA, Demanda.Status.CANCELADA]
    ).count()
    
    # Estatísticas por prioridade
    prioridade_stats = {
        i: Demanda.objects.filter(prioridade=i).count()
        for i in range(1, 6)
    }
    
    # Alertas - demandas com prazo próximo
    data_limite = hoje + timedelta(days=DAYS_ALERT_THRESHOLD)
    demandas_alerta = Demanda.objects.filter(
        data_prazo__lte=data_limite,
        data_prazo__gte=hoje
    ).exclude(
        status__in=[Demanda.Status.CONCLUIDA, Demanda.Status.CANCELADA]
    ).order_by('data_prazo')[:10]
    
    # Métricas de desempenho
    tempo_medio_conclusao = _calcular_tempo_medio_conclusao()
    taxa_no_prazo = _calcular_taxa_no_prazo()
    
    # Distribuições
    projetos_stats = Demanda.objects.values('projeto').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    top_solicitantes = Demanda.objects.values('solicitante').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    top_responsaveis = Demanda.objects.exclude(
        responsavel=''
    ).values('responsavel').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    context = {
        # Contagens por status
        'total_demandas': Demanda.objects.count(),
        **status_counts,
        'demandas_atrasadas': demandas_atrasadas,
        
        # Estatísticas
        'prioridade_stats': prioridade_stats,
        'demandas_alerta': demandas_alerta,
        'demandas_recentes': Demanda.objects.order_by('-criado_em')[:5],
        'minhas_demandas': Demanda.objects.filter(criado_por=request.user).count(),
        
        # Dados para gráficos
        'status_data': _get_status_chart_data(status_counts),
        'notificacoes': get_notificacoes_prazo(),
        
        # Métricas de desempenho
        'tempo_medio_conclusao': round(tempo_medio_conclusao, 1),
        'taxa_no_prazo': round(taxa_no_prazo, 1),
        'projetos_stats': projetos_stats,
        'top_solicitantes': top_solicitantes,
        'top_responsaveis': top_responsaveis,
    }
    
    return render(request, 'demandas/dashboard.html', context)


def _get_status_counts() -> dict:
    """Retorna contagem de demandas por status."""
    return {
        'demandas_pendentes': Demanda.objects.filter(status=Demanda.Status.PENDENTE).count(),
        'demandas_andamento': Demanda.objects.filter(status=Demanda.Status.ANDAMENTO).count(),
        'demandas_concluidas': Demanda.objects.filter(status=Demanda.Status.CONCLUIDA).count(),
        'demandas_canceladas': Demanda.objects.filter(status=Demanda.Status.CANCELADA).count(),
        'demandas_pausa': Demanda.objects.filter(status=Demanda.Status.PAUSA).count(),
    }


def _get_status_chart_data(status_counts: dict) -> dict:
    """Retorna dados formatados para gráfico de status."""
    return {
        'labels': ['Pendente', 'Em Andamento', 'Concluída', 'Cancelada', 'Em Pausa'],
        'data': [
            status_counts['demandas_pendentes'],
            status_counts['demandas_andamento'],
            status_counts['demandas_concluidas'],
            status_counts['demandas_canceladas'],
            status_counts['demandas_pausa'],
        ],
        'colors': ['#ffc107', '#17a2b8', '#28a745', '#dc3545', '#6c757d']
    }


def _calcular_tempo_medio_conclusao() -> float:
    """Calcula tempo médio de conclusão em dias."""
    demandas = Demanda.objects.filter(
        status=Demanda.Status.CONCLUIDA,
        data_conclusao__isnull=False,
        data_entrada__isnull=False
    )
    
    if not demandas.exists():
        return 0.0
    
    total_dias = sum(
        (d.data_conclusao - d.data_entrada).days
        for d in demandas
    )
    return total_dias / demandas.count()


def _calcular_taxa_no_prazo() -> float:
    """Calcula percentual de demandas concluídas dentro do prazo."""
    demandas_com_prazo = Demanda.objects.filter(
        status=Demanda.Status.CONCLUIDA,
        data_conclusao__isnull=False,
        data_prazo__isnull=False
    )
    
    if not demandas_com_prazo.exists():
        return 0.0
    
    demandas_no_prazo = demandas_com_prazo.filter(
        data_conclusao__lte=F('data_prazo')
    ).count()
    
    return (demandas_no_prazo / demandas_com_prazo.count()) * 100


# =============================================================================
# CRUD DE DEMANDAS
# =============================================================================


class DemandaListView(LoginRequiredMixin, ListView):
    """
    View de listagem de demandas com busca e filtros avançados.
    
    Suporta filtros por: status, prioridade, projeto, responsável,
    tag e período de data. Inclui ordenação e paginação.
    """
    
    model = Demanda
    template_name = 'demandas/demanda_list.html'
    context_object_name = 'demandas'
    paginate_by = DEMANDAS_PER_PAGE
    
    VALID_ORDER_FIELDS = {
        'data_prazo', '-data_prazo',
        'prioridade', '-prioridade',
        'status', '-status',
        'criado_em', '-criado_em',
    }
    
    def get_queryset(self):
        """Aplica filtros e ordenação ao queryset."""
        queryset = Demanda.objects.all()
        queryset = self._apply_search_filter(queryset)
        queryset = self._apply_field_filters(queryset)
        queryset = self._apply_date_filters(queryset)
        queryset = self._apply_ordering(queryset)
        return queryset.distinct()
    
    def _apply_search_filter(self, queryset):
        """Aplica filtro de busca textual."""
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(titulo__icontains=search) |
                Q(descricao__icontains=search) |
                Q(codigo__icontains=search)
            )
        return queryset
    
    def _apply_field_filters(self, queryset):
        """Aplica filtros por campos específicos."""
        filters = {
            'status': 'status',
            'prioridade': 'prioridade',
            'projeto': 'projeto__icontains',
            'responsavel': 'responsavel__icontains',
            'tag': 'tags__id',
        }
        
        for param, field in filters.items():
            value = self.request.GET.get(param)
            if value:
                queryset = queryset.filter(**{field: value})
        
        return queryset
    
    def _apply_date_filters(self, queryset):
        """Aplica filtros por período de data."""
        data_inicio = self.request.GET.get('data_inicio')
        data_fim = self.request.GET.get('data_fim')
        
        if data_inicio:
            queryset = queryset.filter(data_entrada__gte=data_inicio)
        if data_fim:
            queryset = queryset.filter(data_entrada__lte=data_fim)
        
        return queryset
    
    def _apply_ordering(self, queryset):
        """Aplica ordenação ao queryset."""
        order_by = self.request.GET.get('order_by', '-criado_em')
        if order_by in self.VALID_ORDER_FIELDS:
            queryset = queryset.order_by(order_by)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tags'] = Tag.objects.all()
        context['status_choices'] = Demanda.STATUS_CHOICES
        context['prioridade_choices'] = Demanda.PRIORIDADE_CHOICES
        
        # Manter filtros nos links de paginação
        get_copy = self.request.GET.copy()
        if 'page' in get_copy:
            get_copy.pop('page')
        context['query_params'] = get_copy.urlencode()
        
        return context


class DemandaDetailView(LoginRequiredMixin, DetailView):
    """Visualização detalhada de uma demanda"""
    model = Demanda
    template_name = 'demandas/demanda_detail.html'
    context_object_name = 'demanda'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['comentarios'] = self.object.comentarios.all().order_by('-criado_em')
        context['historico'] = self.object.historico.all().order_by('-data_alteracao')
        context['anexos'] = self.object.anexos.all().order_by('-enviado_em')
        context['comentario_form'] = ComentarioForm()
        context['anexo_form'] = AnexoForm()
        return context


class DemandaCreateView(LoginRequiredMixin, CreateView):
    """Criação de nova demanda"""
    model = Demanda
    form_class = DemandaForm
    template_name = 'demandas/demanda_form.html'
    
    def form_valid(self, form):
        form.instance.criado_por = self.request.user
        messages.success(self.request, 'Demanda criada com sucesso!')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('demandas:demanda_detail', kwargs={'pk': self.object.pk})


class DemandaUpdateView(LoginRequiredMixin, UpdateView):
    """Edição de demanda"""
    model = Demanda
    form_class = DemandaForm
    template_name = 'demandas/demanda_form.html'
    
    def form_valid(self, form):
        messages.success(self.request, 'Demanda atualizada com sucesso!')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('demandas:demanda_detail', kwargs={'pk': self.object.pk})


class DemandaDeleteView(LoginRequiredMixin, DeleteView):
    """Exclusão de demanda com controle de permissões"""
    model = Demanda
    template_name = 'demandas/demanda_confirm_delete.html'
    success_url = reverse_lazy('demandas:demanda_list')
    
    def dispatch(self, request, *args, **kwargs):
        """Verificar permissões antes de permitir acesso"""
        demanda = self.get_object()
        
        # Apenas criador da demanda ou admin pode excluir
        if not (request.user == demanda.solicitante or request.user.is_superuser):
            messages.error(request, 'Você não tem permissão para excluir esta demanda. Apenas o solicitante ou administradores podem fazê-lo.')
            return redirect('demandas:demanda_detail', pk=demanda.pk)
        
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        """Validar confirmação dupla antes de excluir"""
        confirmacao = self.request.POST.get('confirmacao')
        demanda = self.get_object()
        
        if confirmacao != demanda.codigo:
            messages.error(self.request, 'Confirmação incorreta. Digite exatamente o código da demanda para confirmar a exclusão.')
            return self.form_invalid(form)
        
        messages.success(self.request, 'Demanda excluída com sucesso!')
        return super().form_valid(form)


@login_required
def matriz_eisenhower(request):
    """View para exibir demandas na Matriz de Eisenhower"""
    # Quadrante 1: Urgente e Importante (Alta/Crítica + prioridade 4-5)
    q1 = Demanda.objects.filter(
        criticidade__in=['alta', 'critica'],
        prioridade__in=[4, 5]
    ).exclude(status__in=['concluida', 'cancelada'])
    
    # Quadrante 2: Não urgente e Importante (Alta/Crítica + prioridade 1-3)
    q2 = Demanda.objects.filter(
        criticidade__in=['alta', 'critica'],
        prioridade__in=[1, 2, 3]
    ).exclude(status__in=['concluida', 'cancelada'])
    
    # Quadrante 3: Urgente e Não importante (Baixa/Média + prioridade 4-5)
    q3 = Demanda.objects.filter(
        criticidade__in=['baixa', 'media'],
        prioridade__in=[4, 5]
    ).exclude(status__in=['concluida', 'cancelada'])
    
    # Quadrante 4: Não urgente e Não importante (Baixa/Média + prioridade 1-3)
    q4 = Demanda.objects.filter(
        criticidade__in=['baixa', 'media'],
        prioridade__in=[1, 2, 3]
    ).exclude(status__in=['concluida', 'cancelada'])
    
    context = {
        'quadrante1': q1,
        'quadrante2': q2,
        'quadrante3': q3,
        'quadrante4': q4,
    }
    
    return render(request, 'demandas/matriz_eisenhower.html', context)


@login_required
def adicionar_comentario(request, pk):
    """Adicionar comentário a uma demanda (Ajax)"""
    if request.method == 'POST':
        demanda = get_object_or_404(Demanda, pk=pk)
        form = ComentarioForm(request.POST)
        
        if form.is_valid():
            comentario = form.save(commit=False)
            comentario.demanda = demanda
            comentario.usuario = request.user
            comentario.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Comentário adicionado com sucesso!',
                    'comentario': {
                        'usuario': comentario.usuario.get_full_name() or comentario.usuario.username,
                        'texto': comentario.texto,
                        'criado_em': comentario.criado_em.strftime('%d/%m/%Y %H:%M')
                    }
                })
            else:
                messages.success(request, 'Comentário adicionado com sucesso!')
                return redirect('demandas:demanda_detail', pk=pk)
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    
    return redirect('demandas:demanda_detail', pk=pk)


@login_required
def upload_anexo(request, pk):
    """Upload de anexo para uma demanda"""
    if request.method == 'POST':
        demanda = get_object_or_404(Demanda, pk=pk)
        form = AnexoForm(request.POST, request.FILES)
        
        if form.is_valid():
            anexo = form.save(commit=False)
            anexo.demanda = demanda
            anexo.enviado_por = request.user
            anexo.save()
            
            messages.success(request, 'Anexo enviado com sucesso!')
        else:
            messages.error(request, 'Erro ao enviar anexo. Verifique o arquivo.')
    
    return redirect('demandas:demanda_detail', pk=pk)


# EXPORTAÇÃO DE RELATÓRIOS

@login_required
def exportar_excel(request):
    """Exporta lista de demandas para Excel com filtros aplicados"""
    
    # Recuperar filtros da sessão ou da query string
    filtros = {}
    if 'busca' in request.GET:
        filtros['busca'] = request.GET.get('busca', '')
    if 'status' in request.GET:
        filtros['status'] = request.GET.get('status', '')
    if 'prioridade' in request.GET:
        filtros['prioridade'] = request.GET.get('prioridade', '')
    if 'categoria' in request.GET:
        filtros['categoria'] = request.GET.get('categoria', '')
    if 'responsavel' in request.GET:
        filtros['responsavel'] = request.GET.get('responsavel', '')
    
    # Filtrar demandas
    demandas = Demanda.objects.select_related('solicitante', 'responsavel', 'categoria').prefetch_related('tags').all()
    
    if filtros.get('busca'):
        demandas = demandas.filter(
            Q(titulo__icontains=filtros['busca']) |
            Q(descricao__icontains=filtros['busca']) |
            Q(codigo__icontains=filtros['busca'])
        )
    
    if filtros.get('status'):
        demandas = demandas.filter(status=filtros['status'])
    
    if filtros.get('prioridade'):
        demandas = demandas.filter(prioridade=filtros['prioridade'])
    
    if filtros.get('categoria'):
        demandas = demandas.filter(categoria_id=filtros['categoria'])
    
    if filtros.get('responsavel'):
        demandas = demandas.filter(responsavel_id=filtros['responsavel'])
    
    # Criar workbook do Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demandas"
    
    # Cabeçalhos
    headers = [
        'Código', 'Título', 'Descrição', 'Status', 'Prioridade', 'Criticidade',
        'Solicitante', 'Responsável', 'Categoria', 'Projeto', 'Tags',
        'Data Entrada', 'Data Prazo', 'Data Conclusão', 'Observações'
    ]
    
    # Escrever cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Escrever dados
    for row_num, demanda in enumerate(demandas, 2):
        ws.cell(row=row_num, column=1, value=demanda.codigo)
        ws.cell(row=row_num, column=2, value=demanda.titulo)
        ws.cell(row=row_num, column=3, value=demanda.descricao)
        ws.cell(row=row_num, column=4, value=demanda.get_status_display())
        ws.cell(row=row_num, column=5, value=demanda.prioridade)
        ws.cell(row=row_num, column=6, value=demanda.get_criticidade_display())
        ws.cell(row=row_num, column=7, value=str(demanda.solicitante))
        ws.cell(row=row_num, column=8, value=str(demanda.responsavel) if demanda.responsavel else '')
        ws.cell(row=row_num, column=9, value=str(demanda.categoria) if demanda.categoria else '')
        ws.cell(row=row_num, column=10, value=demanda.projeto)
        ws.cell(row=row_num, column=11, value=', '.join([tag.nome for tag in demanda.tags.all()]))
        ws.cell(row=row_num, column=12, value=demanda.data_entrada.strftime('%d/%m/%Y') if demanda.data_entrada else '')
        ws.cell(row=row_num, column=13, value=demanda.data_prazo.strftime('%d/%m/%Y') if demanda.data_prazo else '')
        ws.cell(row=row_num, column=14, value=demanda.data_conclusao.strftime('%d/%m/%Y') if demanda.data_conclusao else '')
        ws.cell(row=row_num, column=15, value=demanda.observacoes)
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Aplicar filtros
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).coordinate}"
    
    # Preparar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"demandas_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def exportar_pdf(request):
    """Exporta lista de demandas para PDF com filtros aplicados"""
    
    # Recuperar filtros da sessão ou da query string
    filtros = {}
    if 'busca' in request.GET:
        filtros['busca'] = request.GET.get('busca', '')
    if 'status' in request.GET:
        filtros['status'] = request.GET.get('status', '')
    if 'prioridade' in request.GET:
        filtros['prioridade'] = request.GET.get('prioridade', '')
    if 'categoria' in request.GET:
        filtros['categoria'] = request.GET.get('categoria', '')
    if 'responsavel' in request.GET:
        filtros['responsavel'] = request.GET.get('responsavel', '')
    
    # Filtrar demandas
    demandas = Demanda.objects.select_related('solicitante', 'responsavel', 'categoria').prefetch_related('tags').all()
    
    if filtros.get('busca'):
        demandas = demandas.filter(
            Q(titulo__icontains=filtros['busca']) |
            Q(descricao__icontains=filtros['busca']) |
            Q(codigo__icontains=filtros['busca'])
        )
    
    if filtros.get('status'):
        demandas = demandas.filter(status=filtros['status'])
    
    if filtros.get('prioridade'):
        demandas = demandas.filter(prioridade=filtros['prioridade'])
    
    if filtros.get('categoria'):
        demandas = demandas.filter(categoria_id=filtros['categoria'])
    
    if filtros.get('responsavel'):
        demandas = demandas.filter(responsavel_id=filtros['responsavel'])
    
    # Criar documento PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f"demandas_{datetime.now().strftime('%Y-%m-%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.alignment = 1  # Centralizado
    
    # Título
    title = Paragraph("Relatório de Demandas", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Informações do relatório
    info_style = styles['Normal']
    info_text = f"""
    <b>Data de Geração:</b> {datetime.now().strftime('%d/%m/%Y às %H:%M')}<br/>
    <b>Total de Demandas:</b> {demandas.count()}<br/>
    """
    
    if any(filtros.values()):
        info_text += "<b>Filtros Aplicados:</b><br/>"
        if filtros.get('busca'):
            info_text += f"• Busca: {filtros['busca']}<br/>"
        if filtros.get('status'):
            info_text += f"• Status: {filtros['status']}<br/>"
        if filtros.get('prioridade'):
            info_text += f"• Prioridade: {filtros['prioridade']}<br/>"
        if filtros.get('categoria'):
            info_text += f"• Categoria: {filtros['categoria']}<br/>"
        if filtros.get('responsavel'):
            info_text += f"• Responsável: {filtros['responsavel']}<br/>"
    
    info_para = Paragraph(info_text, info_style)
    elements.append(info_para)
    elements.append(Spacer(1, 20))
    
    # Tabela de dados
    data = [['Código', 'Título', 'Status', 'Prioridade', 'Responsável', 'Prazo']]
    
    for demanda in demandas:
        data.append([
            demanda.codigo,
            demanda.titulo[:30] + '...' if len(demanda.titulo) > 30 else demanda.titulo,
            demanda.get_status_display(),
            str(demanda.prioridade),
            str(demanda.responsavel)[:20] if demanda.responsavel else '',
            demanda.data_prazo.strftime('%d/%m/%Y') if demanda.data_prazo else ''
        ])
    
    # Criar tabela
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    return response


# SISTEMA DE NOTIFICAÇÕES

def get_notificacoes_prazo():
    """Retorna demandas com prazo próximo (até 7 dias) ou atrasadas"""
    hoje = timezone.now().date()
    prazo_limite = hoje + timedelta(days=7)
    
    # Demandas atrasadas
    atrasadas = Demanda.objects.filter(
        data_prazo__lt=hoje,
        status__in=['pendente', 'em_andamento']
    ).count()
    
    # Demandas com prazo próximo
    prazo_proximo = Demanda.objects.filter(
        data_prazo__gte=hoje,
        data_prazo__lte=prazo_limite,
        status__in=['pendente', 'em_andamento']
    ).count()
    
    return {
        'atrasadas': atrasadas,
        'prazo_proximo': prazo_proximo,
        'total': atrasadas + prazo_proximo
    }


@login_required
def get_notificacoes_json(request):
    """Retorna notificações em formato JSON para AJAX"""
    notificacoes = get_notificacoes_prazo()
    return JsonResponse(notificacoes)


# GESTÃO DE TAGS

@login_required
def tags_list(request):
    """Lista todas as tags"""
    tags = Tag.objects.annotate(
        demandas_count=Count('demandas')
    ).order_by('nome')
    
    context = {
        'tags': tags,
    }
    return render(request, 'demandas/tags_list.html', context)


@login_required
def tag_create(request):
    """Criar nova tag"""
    from .forms import TagForm
    
    if request.method == 'POST':
        form = TagForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tag criada com sucesso!')
            return redirect('demandas:tags_list')
    else:
        form = TagForm()
    
    return render(request, 'demandas/tag_form.html', {'form': form, 'action': 'Criar'})


@login_required
def tag_edit(request, pk):
    """Editar tag"""
    from .forms import TagForm
    tag = get_object_or_404(Tag, pk=pk)
    
    if request.method == 'POST':
        form = TagForm(request.POST, instance=tag)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tag atualizada com sucesso!')
            return redirect('demandas:tags_list')
    else:
        form = TagForm(instance=tag)
    
    return render(request, 'demandas/tag_form.html', {'form': form, 'action': 'Editar', 'tag': tag})


@login_required
def tag_delete(request, pk):
    """Excluir tag"""
    tag = get_object_or_404(Tag, pk=pk)
    
    if request.method == 'POST':
        tag_nome = tag.nome
        tag.delete()
        messages.success(request, f'Tag "{tag_nome}" excluída com sucesso!')
        return redirect('demandas:tags_list')
    
    return render(request, 'demandas/tag_confirm_delete.html', {'tag': tag})


# Função auxiliar para redirecionar para dashboard (compatibilidade)
def index(request):
    """Redirecionamento para dashboard"""
    return redirect('demandas:dashboard')
